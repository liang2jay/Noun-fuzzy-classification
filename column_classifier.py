import openpyxl
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.svm import SVC
import pickle


class Classifier:

    def __init__(self, table_destination, filename):
        self.data = openpyxl.load_workbook('00_training_sample.xlsx')
        self.sheet = self.data.active
        self.list_column = []
        self.list_column_mapping = []
        self.X_test = []
        self.y_test = []
        self.table_destination = table_destination
        self.filename = filename
        self.list_vector_base = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q',
                                 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', '1', '2', '3', '4', '5', '6', '7', '8',
                                 '9', '0']
        self.list_vector_base02 = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q',
                                 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', '1', '2', '3', '4', '5', '6', '7', '8',
                                 '9', '0', 'verrech_belegnr', 'verrech_belegnr', 'verrech_belegnr', 'verrech_belegnr',
                                 'verrech_belegnr', 'verrech_belegnr', 'verrech_belegnr', 'verrech_belegnr',
                                 'verrech_belegnr', 'verrech_belegnr', 'Interne', 'Interne', 'Interne', 'Interne',
                                 'Interne', 'Interne', 'Interne', 'Interne', 'Interne', 'Interne']

    def corpus_extract(self):
        for row in range(1, 557):
            if self.sheet.cell(row=row, column=4).value == self.table_destination:
                self.list_column.append(self.sheet.cell(row=row, column=2).value.lower())
                prefix_temp, column_temp = self.sheet.cell(row=row, column=5).value.split("_", 1)
                self.list_column_mapping.append(column_temp.lower())
        list_column_original = self.list_column
        list_column_mapping_original = self.list_column_mapping
        for num in range(1, 100):
            self.list_column = self.list_column + list_column_original
            self.list_column_mapping = self.list_column_mapping + list_column_mapping_original
        print("The size of list_column and list_column_mapping is",
              len(self.list_column), ",", len(self.list_column_mapping))

    def data_processing_training(self, list_vector_base):
        list_column_vector_criterion = []
        for item in self.list_column:
            list_temp = item.split("_")
            for i in list_temp:
                list_column_vector_criterion.append(i.lower())
        unique_list_table_vector_criterion = list(set(list_column_vector_criterion))
        a_collective_statement_vector_criterion = unique_list_table_vector_criterion + list_vector_base
        a_collective_statement_vector_criterion.sort()
        list_column_turn_to_vector = []
        for item in self.list_column:
            temp = []
            for i in a_collective_statement_vector_criterion:
                num = item.count(i)  # there
                num_weight = num * (len(i))
                temp.append(num_weight)
            list_column_turn_to_vector.append(temp)
        print("---------------------------------------------------------------------------------------------")
        print(a_collective_statement_vector_criterion)
        X_train, self.X_test, y_train, self.y_test = train_test_split(list_column_turn_to_vector,
                                                                      self.list_column_mapping, test_size=0.20)
        classifier_pre = SVC(probability=True)
        classifier_pre.fit(X_train, y_train)
        pickle.dump(classifier_pre, open(self.filename, 'wb'))

    def model_test(self):
        # load the model from disk
        classifier = pickle.load(open(self.filename, 'rb'))
        y_pred = classifier.predict(self.X_test)
        print('-----------------------------------------------------------------')
        print(self.X_test)
        print('-----------------------------------------------------------------')
        print(confusion_matrix(self.y_test, y_pred))
        print('-----------------------------------------------------------------')
        print(classification_report(self.y_test, y_pred))



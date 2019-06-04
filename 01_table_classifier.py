import openpyxl
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.svm import SVC
import pickle


class Classifier:

    def __init__(self):
        self.corpus = openpyxl.load_workbook('00_training_sample.xlsx')
        self.sheet = self.corpus.active
        self.list_table = []
        self.list_table_mapping = []
        self.list_table_vector = []
        self.list_table_vector_criterion_new = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n',
                                                'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', 'journal',
                                                'unternehmen', 'sstamm', 'sammeln', 'nachweis', 'kostenstellen',
                                                'buchungen', 'fibu', 'kosten', 'offene', 'posten', 'gemeinsame',
                                                'daten', 'itor', 'mandanten', 'stamm', 'itoren', 'personen', 'konten',
                                                'mandanten', 'sach', 'kred', 'deb', 'konto', 'adressen',
                                                'adresstabelle', 'adressen', 'adresstabelle', 'adressen',
                                                'adresstabelle']
        self.X_test = []
        self.y_test = []

    def pre_processing(self):

        for row in range(1, 557):
            self.list_table.append(self.sheet.cell(row=row, column=1).value)
            self.list_table_mapping.append(self.sheet.cell(row=row, column=4).value)

        list_table_original = self.list_table
        list_table_mapping_original = self.list_table_mapping
        for num in range(1, 100):
            self.list_table = self.list_table + list_table_original
            self.list_table_mapping = self.list_table_mapping + list_table_mapping_original

        for item in self.list_table:
            temp = []
            for i in self.list_table_vector_criterion_new:
                num = item.count(i)
                num_weight = num * (len(i))
                temp.append(num_weight)
            self.list_table_vector.append(temp)

    def training(self):
        X_train, self.X_test, y_train, self.y_test = train_test_split(self.list_table_vector,
                                                                      self.list_table_mapping, test_size=0.20)
        classifier_pre = SVC(probability=True)
        classifier_pre.fit(X_train, y_train)
        filename = 'table_mapping_model.sav'
        pickle.dump(classifier_pre, open(filename, 'wb'))

    def test_model(self):
        filename = 'table_mapping_model.sav'
        classifier = pickle.load(open(filename, 'rb'))
        y_pred = classifier.predict(self.X_test)
        print('-----------------------------------------------------------------')
        print(self.X_test)
        print('-----------------------------------------------------------------')
        print(confusion_matrix(self.y_test, y_pred))
        print('-----------------------------------------------------------------')
        print(classification_report(self.y_test, y_pred))


if __name__ == "__main__":
    classifier_engine = Classifier()
    classifier_engine.pre_processing()
    classifier_engine.training()
    classifier_engine.test_model()

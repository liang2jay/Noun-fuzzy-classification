import openpyxl
from sklearn.svm import SVC
import pickle
import re


class Classifier:

    def __init__(self, table_source, filename):
        self.data = openpyxl.load_workbook('00_GoBD_structure_complete.xlsx')
        self.sheet = self.data.active
        self.list_source_column = []
        self.list_importance_label = []
        self.table_source = table_source
        self.filename = filename
        self.column_vector_criterion = []
        self.list_column_turn_to_vector = []
        self.source_table = []
        self.list_vector_base = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q',
                                 'r', 's', 't', 'u', 'v', 'w', 'x', 'y', 'z', '1', '2', '3', '4', '5', '6', '7', '8',
                                 '9', '0']

    def corpus_extract(self):
        source_table_temp = []
        for row in range(2, 1531):
            if self.sheet.cell(row=row, column=2).value == self.table_source:
                self.list_source_column.append(self.sheet.cell(row=row, column=3).value.lower())
                if (re.search('[a-z]', str(self.sheet.cell(row=row, column=4).value))) and \
                                                         (re.search('[a-z]', str(self.sheet.cell(row=row, column=5)))):
                    self.list_importance_label.append('1')
                else:
                    self.list_importance_label.append('0')
            source_table_temp.append(self.sheet.cell(row=row, column=2).value)

        self.source_table = sorted(list(set(source_table_temp)))

    def criterion_design(self):
        list_column_vector_criterion = []
        for item in self.list_source_column:
            list_temp = item.split("_")
            for i in list_temp:
                list_column_vector_criterion.append(i.lower())
        unique_list_table_vector_criterion = list(set(list_column_vector_criterion))
        self.column_vector_criterion = unique_list_table_vector_criterion + self.list_vector_base
        self.column_vector_criterion.sort()
        print(self.column_vector_criterion)

    def data_training(self):
        for item in self.list_source_column:
            temp = []
            for i in self.column_vector_criterion:
                num = item.count(i)  # there
                num_weight = num * (len(i))
                temp.append(num_weight)
            self.list_column_turn_to_vector.append(temp)

        list_column_turn_to_vector = self.list_column_turn_to_vector
        list_importance_label = self.list_importance_label
        for num in range(1, 200):
            self.list_column_turn_to_vector = self.list_column_turn_to_vector + list_column_turn_to_vector
            self.list_importance_label = self.list_importance_label + list_importance_label

        X = self.list_column_turn_to_vector
        y = self.list_importance_label
        classifier_svm = SVC(probability=True)
        classifier_svm.fit(X, y)
        pickle.dump(classifier_svm, open(self.filename, 'wb'))


class ModelFactory:

    def __init__(self):
        classifier_iterate = Classifier('none', 'none.sav')
        classifier_iterate.corpus_extract()
        self.source_table_m = classifier_iterate.source_table
        self.special_item_one = []
        self.special_item_zero = []
        self.special_item_zero_and_one = []

    def create_column_model(self):
        for table in self.source_table_m:
            table_model_name = table + '.sav'
            classifier_iterate = Classifier(table, table_model_name)
            classifier_iterate.corpus_extract()
            if len(list(set(classifier_iterate.list_importance_label))) == 2:
                self.special_item_zero_and_one.append(table)
                print('----------------------------------------------------------')
                print(table)
                classifier_iterate.criterion_design()
                print('----------------------------------------------------------')
                classifier_iterate.data_training()
            else:
                if list(set(classifier_iterate.list_importance_label))[0] is '1':
                    self.special_item_one.append(table)
                else:
                    self.special_item_zero.append(table)


if __name__ == '__main__':
    Start = ModelFactory()
    Start.create_column_model()
    """
    print(Start.special_item_one)
    print('------------------------------------------------------------------')
    print(Start.special_item_zero)
    print('------------------------------------------------------------------')
    print(Start.special_item_zero_and_one)
    print(len(Start.special_item_one) + len(Start.special_item_zero) + len(Start.special_item_zero_and_one))
    """






















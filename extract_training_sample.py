import openpyxl
from openpyxl import Workbook


class Extraction:

    def __init__(self):
        self.data_source = openpyxl.load_workbook('00_GoBD_structure_complete.xlsx')
        self.sheet = self.data_source.active
        self.list_table = []
        self.list_column = []
        self.list_data_type = []
        self.list_table_mapping = []
        self.list_column_mapping = []
        self.list_final = []

    def extra(self):
        for row in range(2, 1531):
            self.list_table.append(self.sheet.cell(row=row, column=2).value)
            self.list_column.append(self.sheet.cell(row=row, column=3).value)
            self.list_table_mapping.append(self.sheet.cell(row=row, column=4).value)
            self.list_column_mapping.append(self.sheet.cell(row=row, column=5).value)

        list_table_save = []
        list_column_save = []
        list_data_type_save = []
        list_table_mapping_save = []
        list_column_mapping_save = []

        for item in range(len(self.list_table)):
            if self.list_table_mapping[item] != '-' and self.list_column_mapping[item] != \
                    '-' and self.list_column_mapping[item] is not None and self.list_table_mapping[item] is not None:
                list_table_save.append(self.list_table[item])
                list_column_save.append(self.list_column[item])
                list_table_mapping_save.append(self.list_table_mapping[item])
                list_column_mapping_save.append(self.list_column_mapping[item])

        book = Workbook()
        sheet_extract = book.active

        for item_output in range(len(list_table_save)):
            sheet_extract.cell(row=item_output + 1, column=1).value = list_table_save[item_output]
            sheet_extract.cell(row=item_output + 1, column=2).value = list_column_save[item_output]
            sheet_extract.cell(row=item_output + 1, column=4).value = list_table_mapping_save[item_output]
            sheet_extract.cell(row=item_output + 1, column=5).value = list_column_mapping_save[item_output]

        book.save("00_training_sample.xlsx")


if __name__ == '__main__':
    Start = Extraction()
    Start.extra()


